import argparse
import os
import sys
import openpyxl
import pyodbc

""" -------------------------------
  | I N I T I A L I Z A T I O N    |
    -------------------------------
"""
parser = argparse.ArgumentParser(description='Remove Work Orders off LAGAN Tracker Upon Resolution', prefix_chars='-+/')
parser.add_argument('--pswd', action='store', dest='usrpwd', help='Enter database login password')

userdomain = os.getenv('USERDOMAIN')
username = os.getenv('USERNAME')
userprof = os.getenv('USERPROFILE')
# Get the input file
inptBillingfl = 'P02 WO Completions Tracker.xlsx'
laganDir = 'K:\Projects\Work Order n Resource Mgmnt Sys'
inptPath = os.path.join(laganDir, inptBillingfl)
# Database access setup
tserver = 'PWEAMHANSQL2'
#pserver = 'P6PO74SQL01'
tdatabase = 'InforProdTest'
#pdatabase = 'InforProd'
drvr = '{ODBC Driver 13 for SQL Server};'
tuser = userdomain + '\\' + username
sqlqueryf = 'Lagan_query.SQL'
sqlqPath = os.path.join(laganDir, sqlqueryf)
args = parser.parse_args()  # Get command line arguments
tusrpwd = args.usrpwd

""" -------------------------------
    | F U N C T I O N: run_sql_file  |
    -------------------------------
"""
def run_sql_query(filename, connection):
    # ------------------------------------------------
    # The function takes a filename and a connection   |
    # as input and will run the SQL query on the given |
    # connection                                       |
    # ------------------------------------------------
    try:
        inpt_query_f = open(filename)
    except:
        print('File: ' + sqlqueryf + ' cannnot be opened')
    sql = inpt_query_f.read()
    inpt_query_f.close()
    db_cursor = connection.cursor()  # Create a cursor from the connection
    db_cursor.execute(sql)
    row = db_cursor.fetchone()
    mesd001_count = 0
    while row:
        try:
            mesd001_count = mesd001_count + 1
            row = db_cursor.fetchone()
        except TypeError as inst:
            print(inst)
            print(row)
    try:
        print(mesd001_count, 'row(s) selected')
        #ofile.close()
    except:
        print("Output file close error in run_sql_file:", sys.exc_info()[0])
        print('Program ending')
        raise

""" -----------------------------------
  | F U N C T I O N: print_rows_of_data |
    -----------------------------------
"""
def print_rows_of_data():
    # ----------------------------------------------------
    # The function prints the contents of the sheet by row.|
    # ----------------------------------------------------
    book = openpyxl.load_workbook(inptPath)
    sheet = book.active
    cellfmt = "%m/%d/%Y"

    for row in sheet.iter_rows(min_row=3, max_col=11, max_row=sheet.max_row):   #Print rows of data
        for cell in row:
            if cell.col_idx == 1 and cell.column == 'A':
                if cell.value == '=IF($A$2="","",$A$2)':
                    cell.value = sheet['A2'].value.strftime(cellfmt)
                    print(cell.value, end=" ")
                else:
                    print(cell.value, end=" ")
            else:
                print(cell.value, end=" ")
        print()


""" -------------------------------
  | F U N C T I O N: main          |
    -------------------------------
"""
def main():
    #print_rows_of_data()
    try:
        conn = pyodbc.connect(DRIVER=drvr, SERVER=tserver, DATABASE=tdatabase, UID=tuser, PWD=tusrpwd,
                              trusted_connection='yes')
        run_sql_query(sqlqPath, conn)
    except:
        print("Unexpected error in main:", sys.exc_info()[0])
        raise
    finally:
        conn.close
    print('Done')


""" -------------------------------
  | M A I N   P R O G R A M        |
    -------------------------------
"""
if __name__ == "__main__":
    main()
